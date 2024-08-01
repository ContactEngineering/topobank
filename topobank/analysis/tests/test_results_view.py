"""
Test for results view.
"""

import datetime
import tempfile

import ContactMechanics
import muFFT
import NuMPI
import numpy as np
import openpyxl
import pytest
import SurfaceTopography
from django.urls import reverse

import topobank

from ...manager.models import Surface, Topography
from ...manager.tests.utils import SurfaceFactory, Topography1DFactory, UserFactory
from ...manager.utils import subjects_to_base64
from ...utils import assert_in_content
from ..models import Analysis, AnalysisFunction
from ..tasks import current_configuration, perform_analysis
from .utils import SurfaceAnalysisFactory, TopographyAnalysisFactory


def selection_from_instances(instances):
    """A little helper for constructing a selection."""
    x = []
    for i in instances:
        prefix = i.__class__.__name__.lower()
        x.append("{}-{}".format(prefix, i.id))

    return x


def test_selection_from_instances(mocker):
    mocker.patch("topobank.manager.models.Topography")
    mocker.patch("topobank.manager.models.Surface")

    instances = [
        Surface(id=1, name="S1"),
        Surface(id=2, name="S2"),
        Topography(id=1, name="T1"),
        Topography(id=2, name="T2"),
        Topography(id=3, name="T3"),
    ]

    expected = [
        "surface-1",
        "surface-2",
        "topography-1",
        "topography-2",
        "topography-3",
    ]

    assert expected == selection_from_instances(instances)


@pytest.mark.django_db
def test_analysis_times(
    api_client, two_topos, test_analysis_function, handle_usage_statistics
):
    topo = Topography.objects.first()

    # we make sure to have to right user who has access
    user = topo.surface.creator
    api_client.force_login(user)

    analysis = TopographyAnalysisFactory.create(
        subject_topography=topo,
        function=test_analysis_function,
        task_state=Analysis.SUCCESS,
        start_time=datetime.datetime(2018, 1, 1, 12),
        end_time=datetime.datetime(
            2018, 1, 1, 13, 1, 1
        ),  # duration: 1 hour, 1 minute, 1 sec
    )
    analysis.users.add(user)
    analysis.save()

    response = api_client.get(
        reverse(
            "analysis:card-series", kwargs=dict(function_id=test_analysis_function.id)
        )
        + "?subjects="
        + subjects_to_base64([topo])
    )

    assert response.status_code == 200

    analyses = response.data["analyses"]
    assert len(analyses) == 1
    assert analyses[0]["start_time"] == "2018-01-01T12:00:00+01:00"
    assert analyses[0]["duration"] == datetime.timedelta(seconds=3661)


@pytest.mark.django_db
def test_show_only_last_analysis(
    api_client, two_topos, test_analysis_function, handle_usage_statistics
):
    topo1 = Topography.objects.first()
    topo2 = Topography.objects.last()

    user = topo1.surface.creator
    api_client.force_login(user)

    result = {
        "name": "test function",
        "xlabel": "x",
        "ylabel": "y",
        "xunit": "1",
        "yunit": "1",
        "series": [],
    }

    #
    # Topography 1
    #
    analysis = TopographyAnalysisFactory.create(
        subject_topography=topo1,
        function=test_analysis_function,
        task_state=Analysis.SUCCESS,
        kwargs={},
        start_time=datetime.datetime(2018, 1, 1, 12),
        end_time=datetime.datetime(2018, 1, 1, 13, 1, 1),
        result=result,
    )
    analysis.users.add(user)
    analysis.save()

    # save a second only, which has a later start time
    analysis = TopographyAnalysisFactory.create(
        subject_topography=topo1,
        function=test_analysis_function,
        task_state=Analysis.SUCCESS,
        kwargs={},
        start_time=datetime.datetime(2018, 1, 2, 12),
        end_time=datetime.datetime(2018, 1, 2, 13, 1, 1),
        result=result,
    )
    analysis.users.add(user)
    analysis.save()

    #
    # Topography 2
    #
    analysis = TopographyAnalysisFactory.create(
        subject_topography=topo2,
        function=test_analysis_function,
        task_state=Analysis.SUCCESS,
        kwargs={},
        start_time=datetime.datetime(2018, 1, 3, 12),
        end_time=datetime.datetime(2018, 1, 3, 13, 1, 1),
        result=result,
    )
    analysis.users.add(user)
    analysis.save()

    # save a second only, which has a later start time
    analysis = TopographyAnalysisFactory.create(
        subject_topography=topo2,
        function=test_analysis_function,
        task_state=Analysis.SUCCESS,
        kwargs={},
        start_time=datetime.datetime(2018, 1, 4, 12),
        end_time=datetime.datetime(2018, 1, 4, 13, 1, 1),
        result=result,
    )
    analysis.users.add(user)
    analysis.save()

    #
    # Check response, for both topographies only the
    # latest results should be shown
    #
    response = api_client.get(
        reverse(
            "analysis:card-series", kwargs=dict(function_id=test_analysis_function.id)
        )
        + "?subjects="
        + subjects_to_base64([topo1, topo2])
    )

    assert response.status_code == 200

    analyses = response.data["analyses"]
    assert len(analyses) == 2
    assert analyses[0]["start_time"] == "2018-01-02T12:00:00+01:00"
    assert analyses[1]["start_time"] == "2018-01-04T12:00:00+01:00"


@pytest.mark.django_db
def test_warnings_for_different_arguments(api_client, handle_usage_statistics):
    user = UserFactory()
    surf1 = SurfaceFactory(creator=user)
    surf2 = SurfaceFactory(creator=user)
    topo1a = Topography1DFactory(surface=surf1)
    topo1b = Topography1DFactory(surface=surf1)
    topo2a = Topography1DFactory(surface=surf2)

    func = AnalysisFunction.objects.get(name="test")

    #
    # Generate analyses for topographies with differing arguments
    #
    kwargs_1a = dict(a=1, b=2)
    kwargs_1b = dict(a=1, b=3)  # differing from kwargs_1a!
    TopographyAnalysisFactory(
        subject_topography=topo1a, function=func, kwargs=kwargs_1a
    )
    TopographyAnalysisFactory(
        subject_topography=topo1b, function=func, kwargs=kwargs_1b
    )
    TopographyAnalysisFactory(
        subject_topography=topo2a, function=func
    )  # default arguments

    #
    # Generate analyses for surfaces with differing arguments
    #
    kwargs_1 = dict(a=2, b=2)
    kwargs_2 = dict(a=2, b=3)  # differing from kwargs_1a!
    SurfaceAnalysisFactory(subject_surface=surf1, function=func, kwargs=kwargs_1)
    SurfaceAnalysisFactory(subject_surface=surf2, function=func, kwargs=kwargs_2)

    api_client.force_login(user)

    #
    # request card, there should be warnings, one for topographies and one for surfaces
    #
    response = api_client.get(
        reverse("analysis:card-series", kwargs=dict(function_id=func.id))
        + "?subjects="
        + subjects_to_base64([topo1a, topo1b, topo2a, surf1, surf2])
    )

    assert response.status_code == 200
    assert response.data["hasNonuniqueKwargs"]


# Maybe the following test can be rewritten as an integration test for usage with selenium
@pytest.mark.skip("Test makes no sense, because it needs AJAX call to be executed.")
@pytest.mark.django_db
def test_show_multiple_analyses_for_two_functions(client, two_topos):
    username = "testuser"
    password = "abcd$1234"

    assert client.login(username=username, password=password)

    topo1 = Topography.objects.first()
    topo2 = Topography.objects.last()
    af1 = AnalysisFunction.objects.first()
    af2 = AnalysisFunction.objects.last()

    assert topo1 != topo2
    assert af1 != af2

    #
    # Create analyses for two functions and two different topographies
    #
    counter = 0
    for af in [af1, af2]:
        for topo in [topo1, topo2]:
            counter += 1
            analysis = TopographyAnalysisFactory.create(
                subject_topography=topo,
                function=af,
                task_state=Analysis.SUCCESS,
                kwargs={"bins": 10},
                start_time=datetime.datetime(2018, 1, 1, counter),
                end_time=datetime.datetime(2018, 1, 1, counter + 1),
            )
            analysis.save()

    #
    # Select both topographies
    #
    client.post(reverse("manager:topography-select", kwargs=dict(pk=topo1.pk)))
    client.post(reverse("manager:topography-select", kwargs=dict(pk=topo2.pk)))

    #
    # Check response when selecting only first function, both analyses should be shown
    #
    response = client.post(
        reverse("analysis:results-list"),
        data={
            "functions": [af1.id],
        },
        follow=True,
    )

    assert response.status_code == 200

    assert_in_content(response, "Example 3 - ZSensor")
    assert_in_content(response, "Example 4 - Default")

    #
    # Check response when selecting only both functions, both analyses should be shown
    #
    response = client.post(
        reverse("analysis:results-list"),
        data={
            "functions": [af1.id, af2.id],
        },
        follow=True,
    )

    assert response.status_code == 200

    assert_in_content(response, "Example 3 - ZSensor")
    assert_in_content(response, "Example 4 - Default")


@pytest.fixture
def ids_downloadable_analyses(two_topos, settings, test_analysis_function, mocker):
    """Returns ids of analyses which can be downloaded as list."""
    config = current_configuration()

    settings.CELERY_TASK_ALWAYS_EAGER = True  # perform tasks locally

    #
    # create two analyses with results
    #
    topos = [
        Topography.objects.get(name="Example 3 - ZSensor"),
        Topography.objects.get(name="Example 4 - Default"),
    ]

    v = np.arange(5)
    ids = []

    for k in range(2):
        result = dict(
            name=f"Test Results {k}",
            scalars=dict(
                nice_value=13,
                bad_value=-99,
            ),
            xlabel="time",
            ylabel="distance",
            xunit="s",
            yunit="m",
            series=[
                dict(
                    name="First Series",
                    x=v + k,
                    y=2 * v + k,
                ),
                dict(
                    name="Second Series",
                    x=v + 1 + k,
                    y=3 * (v + 1) + k,
                ),
            ],
        )

        analysis = TopographyAnalysisFactory.create(
            subject_topography=topos[k],
            function=test_analysis_function,
            kwargs={},
            configuration=config,
        )

        # we insert our result instead of the real function's result
        m = mocker.patch("topobank.analysis.models.AnalysisFunction.eval")
        m.return_value = result

        # Saving above results in storage
        perform_analysis(analysis.id)

        ids.append(analysis.id)

    return ids


@pytest.mark.django_db
def test_analysis_download_as_txt(
    client, two_topos, ids_downloadable_analyses, settings, handle_usage_statistics
):
    user = two_topos[0].surface.creator  # we need a user which is allowed to download
    client.force_login(user)

    ids_str = ",".join(str(i) for i in ids_downloadable_analyses)
    download_url = reverse(
        "analysis:download", kwargs=dict(ids=ids_str, file_format="txt")
    )

    response = client.get(download_url)

    assert response.status_code == 200

    from io import StringIO

    txt = response.content.decode()

    assert "test" in txt  # function name should be in there

    # check whether version numbers are in there
    assert SurfaceTopography.__version__.split("+")[0] in txt
    assert ContactMechanics.__version__ in txt
    assert NuMPI.__version__ in txt
    assert muFFT.version.description() in txt
    assert topobank.__version__ in txt

    # check whether creator of topography is listed
    topo1, topo2 = two_topos

    assert "Creator" in txt
    assert topo1.creator.name in txt
    assert topo1.creator.orcid_id in txt
    assert topo2.creator.name in txt
    assert topo2.creator.orcid_id in txt

    # remove comments and empty lines
    filtered_lines = []
    for line in txt.splitlines():
        line = line.strip()
        if not line.startswith("#") and len(line) > 0:
            filtered_lines.append(line)
    filtered_txt = "\n".join(filtered_lines)

    arr = np.loadtxt(StringIO(filtered_txt))

    expected_arr = np.array(
        [
            (0, 0),
            (1, 2),
            (2, 4),
            (3, 6),
            (4, 8),
            (1, 3),
            (2, 6),
            (3, 9),
            (4, 12),
            (5, 15),
            (1, 1),
            (2, 3),
            (3, 5),
            (4, 7),
            (5, 9),
            (2, 4),
            (3, 7),
            (4, 10),
            (5, 13),
            (6, 16),
        ]
    )

    assert arr == pytest.approx(expected_arr)


@pytest.mark.parametrize("same_names", [False, True])
@pytest.mark.django_db
def test_analysis_download_as_xlsx(
    client,
    two_topos,
    ids_downloadable_analyses,
    same_names,
    settings,
    handle_usage_statistics,
):
    topos = Topography.objects.all()
    assert len(topos) == 2

    # if tested with "same_names=True", make sure both topographies have the same name
    if same_names:
        topos[0].name = topos[1].name
        topos[0].save()

    first_topo_name = topos[0].name
    second_topo_name = topos[1].name

    first_topo_name_in_sheet_name = first_topo_name
    second_topo_name_in_sheet_name = second_topo_name
    if same_names:
        first_topo_name_in_sheet_name += " (1)"
        second_topo_name_in_sheet_name += " (2)"

    username = "testuser"
    password = "abcd$1234"

    assert client.login(username=username, password=password)

    ids_str = ",".join(str(i) for i in ids_downloadable_analyses)
    download_url = reverse(
        "analysis:download", kwargs=dict(ids=ids_str, file_format="xlsx")
    )

    response = client.get(download_url)

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx")  # will be deleted automatically
    tmp.write(response.content)
    tmp.seek(0)

    xlsx = openpyxl.load_workbook(tmp.name)

    assert (
        len(xlsx.worksheets) == 1 + 1 + 2 * 2
    )  # INDEX, META DATA, 2*2 analysis sheets

    def assert_data_equal(sheet, exp_data):
        first_data_row = 9
        last_data_col = 2
        data = np.array(list(sheet.values))[first_data_row:, : last_data_col + 1]
        np.testing.assert_equal(data, exp_data)

    assert_data_equal(
        xlsx.get_sheet_by_name("analysis-0-series-0"),
        [
            (None, "time (s)", "distance (m)"),
            (0, 0, 0),
            (1, 1, 2),
            (2, 2, 4),
            (3, 3, 6),
            (4, 4, 8),
        ],
    )

    assert_data_equal(
        xlsx.get_sheet_by_name("analysis-0-series-1"),
        [
            (None, "time (s)", "distance (m)"),
            (0, 1, 3),
            (1, 2, 6),
            (2, 3, 9),
            (3, 4, 12),
            (4, 5, 15),
        ],
    )

    assert_data_equal(
        xlsx.get_sheet_by_name("analysis-0-series-1"),
        [
            (None, "time (s)", "distance (m)"),
            (0, 1, 3),
            (1, 2, 6),
            (2, 3, 9),
            (3, 4, 12),
            (4, 5, 15),
        ],
    )

    assert_data_equal(
        xlsx.get_sheet_by_name("analysis-1-series-0"),
        [
            (None, "time (s)", "distance (m)"),
            (0, 1, 1),
            (1, 2, 3),
            (2, 3, 5),
            (3, 4, 7),
            (4, 5, 9),
        ],
    )

    assert_data_equal(
        xlsx.get_sheet_by_name("analysis-1-series-1"),
        [
            (None, "time (s)", "distance (m)"),
            (0, 2, 4),
            (1, 3, 7),
            (2, 4, 10),
            (3, 5, 13),
            (4, 6, 16),
        ],
    )

    # check whether version numbers are available in INFORMATION sheet
    ws = xlsx.get_sheet_by_name("INFORMATION")

    vals = list(ws.values)

    def assert_version_in_vals(modname, version):
        assert (f"Version of '{modname}'", version) in vals

    assert_version_in_vals(
        "SurfaceTopography", SurfaceTopography.__version__.split("+")[0]
    )
    assert_version_in_vals("ContactMechanics", ContactMechanics.__version__)
    assert_version_in_vals("NuMPI", NuMPI.__version__)
    assert_version_in_vals("muFFT", muFFT.version.description())
    assert_version_in_vals("topobank", topobank.__version__)

    # subject names should also be included, as well as the creator
    for t in topos:
        assert ("Subject type", "topography") in vals
        assert ("Subject name", t.name) in vals
        assert ("Creator", str(t.creator)) in vals

    # Check links on INDEX sheet
    ws = xlsx.get_sheet_by_name("INDEX")

    function_name = Analysis.objects.get(id=ids_downloadable_analyses[0]).function.name

    assert list(ws.values) == [
        (
            "Subject name",
            "Subject type",
            "Function name",
            "Data series",
            "Link",
            "Creator",
            "Instrument name",
            "Instrument type",
            "Instrument parameters",
        ),
        (
            first_topo_name,
            "measurement",
            function_name,
            "First Series",
            "Click to jump to sheet 'analysis-0-series-0'",
            str(topos[0].creator),
            None,
            "undefined",
            "{}",
        ),
        (
            first_topo_name,
            "measurement",
            function_name,
            "Second Series",
            "Click to jump to sheet 'analysis-0-series-1'",
            str(topos[0].creator),
            None,
            "undefined",
            "{}",
        ),
        (
            second_topo_name,
            "measurement",
            function_name,
            "First Series",
            "Click to jump to sheet 'analysis-1-series-0'",
            str(topos[1].creator),
            None,
            "undefined",
            "{}",
        ),
        (
            second_topo_name,
            "measurement",
            function_name,
            "Second Series",
            "Click to jump to sheet 'analysis-1-series-1'",
            str(topos[1].creator),
            None,
            "undefined",
            "{}",
        ),
    ]

    first_link_cell = ws["E2"]
    assert first_link_cell.style == "Hyperlink"
    from openpyxl.worksheet.hyperlink import Hyperlink

    assert isinstance(first_link_cell.hyperlink, Hyperlink)


@pytest.mark.django_db
def test_analysis_download_as_xlsx_despite_slash_in_sheetname(
    client,
    two_topos,
    ids_downloadable_analyses,
    django_user_model,
    handle_usage_statistics,
):
    topos = Topography.objects.all()
    assert len(topos) == 2

    # change topography name such that a slash is in there
    topos[0].name = "This is a slash: /"
    topos[0].save()

    #
    # download analysis data as xlsx, the slash should be replaced by "div"
    #
    user = django_user_model.objects.get(username="testuser")
    client.force_login(user)

    ids_str = ",".join(str(i) for i in ids_downloadable_analyses)
    download_url = reverse(
        "analysis:download", kwargs=dict(ids=ids_str, file_format="xlsx")
    )

    response = client.get(download_url)

    # if this works without error, it's okay, let's test also if the sheet can be loaded:

    tmp = tempfile.NamedTemporaryFile(suffix=".xlsx")  # will be deleted automatically
    tmp.write(response.content)
    tmp.seek(0)

    xlsx = openpyxl.load_workbook(tmp.name)

    assert len(xlsx.worksheets) == 1 + 1 + 2 * 2


@pytest.mark.django_db
def test_download_analysis_results_without_permission(
    client,
    two_topos,
    ids_downloadable_analyses,
    django_user_model,
    handle_usage_statistics,
):
    # two_topos belong to a user "testuser"
    user_2 = django_user_model.objects.create_user(username="attacker")
    client.force_login(user_2)

    ids_str = ",".join(str(i) for i in ids_downloadable_analyses)
    download_url = reverse(
        "analysis:download", kwargs=dict(ids=ids_str, file_format="txt")
    )

    response = client.get(download_url)
    assert response.status_code == 403  # Permission denied

    # when user_2 has view permissions for one topography of both, it's still not okay to download
    two_topos[0].surface.share(user_2)
    response = client.get(download_url)
    assert response.status_code == 403  # Permission denied

    # when user_2 has view permissions for all related surfaces, it's okay to download
    two_topos[1].surface.share(user_2)
    response = client.get(download_url)
    assert response.status_code == 200


@pytest.mark.django_db
def test_view_shared_analysis_results(api_client, handle_usage_statistics):
    password = "abcd$1234"

    #
    # create database objects
    #
    user1 = UserFactory(password=password)
    user2 = UserFactory(password=password)

    surface1 = SurfaceFactory(creator=user1)
    surface2 = SurfaceFactory(creator=user2)

    # create topographies + functions + analyses
    func1 = AnalysisFunction.objects.get(name="test")
    # func2 = AnalysisFunctionFactory()

    # Two topographies for surface1
    topo1a = Topography1DFactory(surface=surface1, name="topo1a")
    topo1b = Topography1DFactory(surface=surface1, name="topo1b")

    # One topography for surface2
    topo2a = Topography1DFactory(surface=surface2, name="topo2a")

    # analyses, differentiate by start time
    TopographyAnalysisFactory(
        subject_topography=topo1a,
        function=func1,
        start_time=datetime.datetime(2019, 1, 1, 12),
    )
    TopographyAnalysisFactory(
        subject_topography=topo1b,
        function=func1,
        start_time=datetime.datetime(2019, 1, 1, 13),
    )
    TopographyAnalysisFactory(
        subject_topography=topo2a,
        function=func1,
        start_time=datetime.datetime(2019, 1, 1, 14),
    )

    # Function should have three analyses, all successful (the default when using the factory)
    assert func1.analysis_set.count() == 3
    assert all(a.task_state == "su" for a in func1.analysis_set.all())

    # user2 shares surfaces, so user 1 should see surface1+surface2
    surface2.share(user1)

    #
    # Now we change to the analysis card view and look what we get
    #
    assert api_client.login(username=user1.username, password=password)

    response = api_client.get(
        reverse("analysis:card-series", kwargs=dict(function_id=func1.id))
        + "?subjects="
        + subjects_to_base64([topo1a, topo1b, topo2a])
    )

    # Function should still have three analyses, all successful (the default when using the factory)
    assert func1.analysis_set.count() == 3
    assert all(a.task_state == "su" for a in func1.analysis_set.all())

    assert response.status_code == 200

    # We should see start times of all three topographies
    analyses = response.data["analyses"]
    assert len(analyses) == 3
    assert analyses[0]["start_time"] == "2019-01-01T12:00:00+01:00"  # topo1a
    assert analyses[1]["start_time"] == "2019-01-01T13:00:00+01:00"  # topo1b
    assert analyses[2]["start_time"] == "2019-01-01T14:00:00+01:00"  # topo2a

    api_client.logout()

    #
    # user 2 cannot access results from topo1, it is not shared
    #
    assert api_client.login(username=user2.username, password=password)

    response = api_client.get(
        reverse("analysis:card-series", kwargs=dict(function_id=func1.id))
        + "?subjects="
        + subjects_to_base64([topo1a, topo1b, topo2a])
    )

    assert response.status_code == 200

    # We should see start times of just one topography
    analyses = response.data["analyses"]
    assert len(analyses) == 1
    assert analyses[0]["start_time"] == "2019-01-01T14:00:00+01:00"  # topo2a

    api_client.logout()
